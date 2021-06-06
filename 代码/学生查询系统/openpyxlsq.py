import openpyxl

wb = openpyxl.load_workbook('income.xlsx')

# Getting sheets for the workbook

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)


# 创建新的单元格
mySheet =  wb.create_sheet('mySheet')
print(wb.sheetnames)

# sheet3 = wb.get_sheet_by_name('Sheet3')
# sheet4 = wb['mySheet']

# Getting cells form get sheets
ws = wb.active
print(ws)
print(ws['A1'])

print(ws['A2'].value)


c = ws['B1']
print('Row {} is {}'.format(c.coordinate,c.value))

print('--------------------------')
print(ws.cell(row=1,column=2).value)

for i in range(1,8,2):
    print(i,ws.cell(row=i,column=2).value)


# 从表单中取行和列

colc = ws['C']
print(colc[2])
print('=====================')
row6 = ws[6]
print(row6)
print('---------------------')
col_range = ws['A:B']
row_range = ws[1:6]
for col in col_range:
    for cell  in col:
        print(cell.value)

print('---------------------')
for row in row_range:
    for cell in row:
        print(cell.value)

for row in ws.iter_cols(min_row=1,max_row=2,max_col=2):
    for cell in row:
        print(cell)

